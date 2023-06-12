from typing import List
import json
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill


class Event:
    def __init__(self, url: str):
        self.url = url
        self.class_name = ""

    @property
    def general_url(self):
        return self.get_general_url(self.event_id)

    @property
    def event_id(self):
        return self.get_event_id()

    @property
    def event_name(self):
        return self.get_event_details(self.general_url)['eventName']

    @property
    def event_start_date(self):
        return self.get_event_details(self.general_url)['startDate']

    @property
    def event_end_date(self):
        return self.get_event_details(self.general_url)['endDate']

    @property
    def class_id(self):
        general_url = self.get_general_url(self.event_id)
        class_ids = self.get_class_ids(general_url)
        return self.get_class_id(class_ids, self.class_name)

    @property
    def class_names(self) -> List[str]:
        return [x[0] for x in self.get_class_ids(self.general_url)]

    @property
    def results(self):
        return self.get_results(self.event_id, self.class_id)

    def get_event_id(self) -> str:
        return self.url[self.url.index("event/") + 6:]

    @staticmethod
    def get_general_url(event_id: str) -> str:
        return f"https://www.manage2sail.com/en-EN/event/{event_id}"

    @staticmethod
    def get_class_ids(html_url: str) -> List:
        classes = []
        ids = []

        web_contents = urllib.request.urlopen(html_url).read()
        soup = BeautifulSoup(web_contents, 'xml')
        results = soup.find(id="classes").find_all("a")

        for _class in soup.find(id="classes").find_all("tr"):
            found_name = _class.find("td")
            if found_name is not None:
                classes.append(_class.find("td").text)

        for url in results:
            if url.find("i").get('title') == "Class has results":
                found_url = url.get('href')
                ids.append(found_url[found_url.index("classId")+8:])

        return list(zip(classes, ids))

    @staticmethod
    def get_event_details(html_url: str):
        details = {}
        web_contents = urllib.request.urlopen(html_url).read()
        soup = BeautifulSoup(web_contents, 'xml')
        details['eventName'] = soup.find("div", {"class": "eventName"}).find("h1").text
        details['startDate'], details['endDate'] = soup.find("span", {"class": "eventDates"}).text.split(" - ")
        return details

    @staticmethod
    def get_class_id(class_ids: List, class_name: str) -> str:
        return class_ids[[x[0] for x in class_ids].index(class_name)][1]

    @staticmethod
    def get_results(event_id: str, class_id: str) -> List:
        contents = urllib.request.urlopen(f"https://www.manage2sail.com/api/event/{event_id}/regattaresult/{class_id}").read()
        result_data = json.loads(contents)
        results = []
        for score in result_data['EntryResults']:
            result = {}
            result['SailNumber'] = score['SailNumber']
            result['Name'] = score['Name']
            result['scores'] = []

            for race in score['EntryRaceResults']:
                discard = False
                if "PointsDiscarded" in race:
                    discard = True
                score = {'points': race['Points'], 'discard': discard}
                result['scores'].append(score)

            results.append(result)
        return results

    def data_dump(self):
        contents = urllib.request.urlopen(f"https://www.manage2sail.com/api/event/{self.event_id}/regattaresult/{self.class_id}").read()
        return json.dumps(json.loads(contents), indent=2)

    def export_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.class_name
        sheet["B1"] = "Sailnumber"
        sheet["C1"] = "Name"
        for index, race in enumerate(self.results[0]['scores']):
            sheet.cell(row=1, column=index+4).value = f'r{index+1}'

        for row, result in enumerate(self.results):
            sheet.cell(row=2 + row, column=1).value = row+1
            sheet.cell(row=2+row, column=2).value = result['SailNumber']
            sheet.cell(row=2 + row, column=3).value = result['Name']
            for race, score in enumerate(result['scores']):
                sheet.cell(row=2 + row, column=4+race).value = float(score['points'])
                if score['discard'] is True:
                    sheet.cell(row=2 + row, column=4 + race).fill = PatternFill("solid", start_color="ff0000")

        workbook.save(filename=f"{self.event_name}.xlsx")


if __name__ == "__main__":
    URL = "https://www.manage2sail.com/nl-NL/event/e5252026-b2c5-4d0a-a077-6bd50d69e55b"

    U4Workum = Event(URL)
    print(U4Workum.event_id)
    print(U4Workum.general_url)
    print(U4Workum.event_name)
    print(U4Workum.event_start_date)
    print(U4Workum.event_end_date)
    print(U4Workum.class_names)
    U4Workum.class_name = 'Optimist'
    U4Workum.export_to_excel()

    print(U4Workum.data_dump())